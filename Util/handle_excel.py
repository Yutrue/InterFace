#coding-utf-8
import openpyxl
import sys
import os
base_path = os.getcwd()
# print(base_path)
sys.path.append(base_path) 

# open_excel = openpyxl.load_workbook(base_path+"/Case/case.xlsx")
# sheet_name = open_excel.sheetnames
# excel_value = open_excel[sheet_name[0]]
# print(excel_value)
# print(excel_value.cell(1,3).value)
# print(excel_value.max_row)

class HandExcle():
    # 加载excel
    def load_excel(self):
        open_excel = openpyxl.load_workbook(base_path+"/Case/imooc.xlsx")
        return open_excel

    #加载sheet
    def get_sheet_data(self,index=None):
        sheet_name = self.load_excel().sheetnames   #加载所有sheets
        if index is None:
            index = 0
        data = self.load_excel()[sheet_name[index]]
        return data

    #获取某一单元格内容
    def get_cell_value(self,row,cols):
        data = self.get_sheet_data().cell(row=row,column=cols).value
        return data

    #获取行数
    def get_rows(self):
        row = self.get_sheet_data().max_row
        return row

    #获取某一行的内容
    def get_rows_value(self,row):
        row_list = []
        for i in self.get_sheet_data()[row]:
            row_list.append(i.value)
        return row_list

    #获取某一列的内容
    def get_cols_value(self,col=None):
        if col is None:
            col='A'
        cols_list = []
        for i in self.get_sheet_data()[col]:
            cols_list.append(i.value)
        return cols_list

    #获取行号
    def get_rows_numble(self,case_id):
        num = 1
        cols_data = self.get_cols_value()
        for col_cata in cols_data:
            if col_cata == case_id:
                return num
            num=num+1
        return num

    #写入数据
    def excel_write_data(self,row,cols,value):
        wb = self.load_excel()
        wr = wb.active
        wr.cell(row,cols,value)
        wb.save(base_path+"/Case/imooc.xlsx")

    #获取excel里所有的数据
    def get_excel_data(self):
        data_list = []
        for i in range(self.get_rows()):
            data_list.append(self.get_rows_value(i+2))

        return data_list

excel_data = HandExcle()

if __name__ == '__main__':
    handle = HandExcle()
    print(handle.get_rows_value(3))
    # handle.excel_write_data(2,12,"通过")
    # print(handle.get_excel_data())